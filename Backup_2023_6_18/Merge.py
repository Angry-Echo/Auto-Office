# -*- coding: utf-8 -*-
import openpyxl
import copy
import glob
import os
import sys
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from Head import Find_Max_Head


def get_max_row(sheet, num_head_row):

    valid_value_position = num_head_row + 1

    while True:
        row_dict = {cell.value for cell in sheet[valid_value_position]}
        if row_dict != {None} and len(row_dict) > 2:
            valid_value_position += 1
        else:
            break

    return valid_value_position - 1


def Merge_Excel(root_dir, head_row, col_id, font_flag):
    # 创建一个新的工作表
    merged = openpyxl.Workbook()
    merged_sheet = merged.active

    # 遍历所有excel文件的sheet,存为list
    sheet_list = []
    try:
        for path in glob.glob(os.path.join(root_dir, '*.xlsx')):  # 搜索给定目录下所有的.xlsx文件
            wb = openpyxl.load_workbook(path)
            sheet = wb.worksheets[0]
            sheet_list.append(sheet)
    except PermissionError:
        print('合并失败，这是因为目录中有Excel文件未关闭！请关闭后重试')
        sys.exit()

    '''
    # 该目录下Excel文件的共同表头的行数
    head_row = Find_Max_Head(sheet_list)
    '''

    # 合并所有sheet中的数据，带格式，复制到新的工作表中
    for n_s, sheet in enumerate(sheet_list):
        num_whole_row = get_max_row(sheet, head_row)

        if n_s == 0:
            place_holder = 0  # 第一个表的表头需要，所有没有行是无效的
            row_begin = 0
        elif n_s == 1:
            place_holder = head_row  # 第二个表的表头不需要，有几行是无效的
            last_max_row = get_max_row(sheet_list[n_s - 1], head_row)  # 现在是第二个表，它的前一个表是第一张表（也就是算入了表头）
            row_begin += last_max_row
        else:
            place_holder = head_row  # 同上
            last_max_row = get_max_row(sheet_list[n_s - 1], head_row) - head_row  # 现在是第三个表，它的前一个表是第二个表（行数不计入表头
            row_begin += last_max_row

        # tab颜色
        merged_sheet.sheet_properties.tabColor = sheet.sheet_properties.tabColor

        # 开始处理合并单元格形式为“(<CellRange A1：A4>,)，替换掉(<CellRange 和 >,)' 找到合并单元格
        wm = list(sheet.merged_cells)
        if len(wm) > 0:
            for i in range(0, len(wm)):
                cell_2 = str(wm[i]).replace('(<CellRange ', '').replace('>,)', '')
                merged_sheet.merge_cells(cell_2)

        for n_r in range(num_whole_row):
            # 设定行高
            merged_sheet.row_dimensions[n_r + 1].height = sheet.row_dimensions[n_r + 1].height

            row = sheet[n_r + place_holder + 1]  # excel 坐标从1开始

            for n_c, source_cell in enumerate(row):
                # 设定列宽
                merged_sheet.column_dimensions[get_column_letter(n_c + 1)].width = sheet.column_dimensions[get_column_letter(n_c + 1)].width

                # 复制单元格，带格式
                target_cell = merged_sheet.cell(row=row_begin + n_r + 1, column=n_c + 1)
                try:
                    target_cell.value = source_cell.value
                except AttributeError:
                    pass  # 合并后只要写第一个即可，后面的都被合并进第一个了，当然找不到了

                target_cell.fill = copy.copy(source_cell.fill)

                if source_cell.has_style:
                    target_cell._style = copy.copy(source_cell._style)
                    target_cell.font = copy.copy(source_cell.font)
                    target_cell.border = copy.copy(source_cell.border)
                    target_cell.fill = copy.copy(source_cell.fill)
                    target_cell.number_format = copy.copy(source_cell.number_format)
                    target_cell.protection = copy.copy(source_cell.protection)
                    target_cell.alignment = copy.copy(source_cell.alignment)

    # 从第二行开始遍历，将 "序号" 列中的值递增
    merged_max_row = get_max_row(merged_sheet, head_row)

    for row in range(head_row + 1, merged_max_row + 1):
        merged_sheet.cell(row=row, column=col_id).value = row - head_row  # 妙！

    # 是否修改字体
    if font_flag == '是':
        for row in range(head_row + 1, merged_max_row + 1):
            sheet_row = merged_sheet[row]
            for cell in sheet_row:
                font = Font(name="仿宋", size=14)
                cell.font = font

    # 创建结算文件夹，保存新的Excel文件
    if not os.path.exists(os.path.join(root_dir, 'Results')):
        os.mkdir(os.path.join(root_dir, 'Results'))

    merged.save(os.path.join(root_dir, 'Results', 'Merged.xlsx'))
    print("合并完成! 汇总项目表保存在: " + os.path.join(root_dir, 'Results', 'Merged.xlsx'))
