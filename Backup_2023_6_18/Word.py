import win32com.client as win32
import openpyxl
import os
import sys
import glob


def Special_Find_Max_Head(root_dir):
    row_flag = 1
    try:
        excel_path = glob.glob(os.path.join(root_dir, '*.xlsx'))

        wb_1 = openpyxl.load_workbook(excel_path[0])
        wb_2 = openpyxl.load_workbook(excel_path[1])

        wb_1_sheet = wb_1.worksheets[0]
        wb_2_sheet = wb_2.worksheets[0]

        while True:
            wb_1_sheet_row = {cell.value for cell in wb_1_sheet[row_flag]}
            wb_2_sheet_row = {cell.value for cell in wb_2_sheet[row_flag]}

            if wb_1_sheet_row == wb_2_sheet_row:
                row_flag += 1
            else:
                break

    except PermissionError:
        print('有Excel文件未关闭！请关闭后重试')
        sys.exit()

    return row_flag - 1  # row_flag比最大行数多一个，是非表头行的位置


def Word_PDF(root_dir, example_path, object_str, porj_col_id, head_row):
    result_dir = os.path.join(root_dir, 'Results')

    # 创建Word应用程序对象
    word = win32.gencache.EnsureDispatch('Word.Application')
    word.Visible = True

    # 打开Excel文件并读取数据
    try:
        wb = openpyxl.load_workbook(os.path.join(result_dir, 'Merged.xlsx'))
        ws = wb.active
    except PermissionError:
        print('打印失败，这是因为合并表Excel文件未关闭！请关闭后重试')
        sys.exit()

    # 从哪一行开始获取项目名称：共同表头的下一行
    proj_start_row = head_row + 1

    # 打开要复制内容的Word文档
    doc1 = word.Documents.Open(example_path)

    # 选择要复制的内容
    selection = word.Selection
    selection.WholeStory()
    selection.Copy()

    # 将内容粘贴到新打开的Word文档中
    doc = word.Documents.Add()
    selection2 = word.Selection
    selection2.EndKey(win32.constants.wdStory)
    selection2.Paste()

    selection2.EndKey(win32.constants.wdStory)
    selection2.InsertBreak(win32.constants.wdPageBreak)

    # 获取文字
    first_cell_value = ws[porj_col_id][proj_start_row - 1].value  # 索引是从0开始，所以 position - 1
    if '《' in first_cell_value:
        pass
    else:
        first_cell_value = '《' + first_cell_value + '》'

    # selection2.Find.Execute('开放性课题', False, False, False, False, False, True, 1, False, first_cell_value, 2)

    # selection2.EndKey(win32.constants.wdStory)
    # selection2.MoveUp(Unit=win32.constants.wdLine, Count=10)
    # selection2.HomeKey(Unit=win32.constants.wdLine)
    # selection2.InsertBefore(first_cell_value)

    # for i in range(10):
    #     selection2.Delete()

    # 查找目标句子
    for sentence in doc.Sentences:
        if object_str in sentence.Text:
            # 找到目标句子后，在其前面插入字符串
            sentence.StartOf(Unit=win32.constants.wdSentence, Extend=win32.constants.wdMove)
            doc.Range(sentence.Start, sentence.Start).InsertBefore(first_cell_value)
            break

    data = [cell.value for cell in ws[porj_col_id][proj_start_row - 1 + 1:] if cell.value is not None]  # 假设需要替换的文本在第一列

    last_target_pos = 0
    for cell_value in data:
        if '《' in cell_value:
            pass
        else:
            cell_value = '《' + cell_value + '》'

        selection2 = word.Selection
        selection2.EndKey(win32.constants.wdStory)
        selection2.Paste()

        selection2.EndKey(win32.constants.wdStory)
        selection2.InsertBreak(win32.constants.wdPageBreak)

        # selection2.Find.Execute('开放性课题', False, False, False, False, False, True, 1, False, cell_value, 2)

        # selection2.EndKey(win32.constants.wdStory)
        # selection2.MoveUp(Unit=win32.constants.wdLine, Count=10)
        # selection2.HomeKey(Unit=win32.constants.wdLine)

        # for i in range(10):
        #     selection2.Delete()

        # 查找目标句子
        # target_text = object_str
        # for sentence in doc.Sentences:
        #     if target_text in sentence.Text:
        #         # 找到目标句子后，在其前面插入字符串
        #         sentence.StartOf(Unit=win32.constants.wdSentence, Extend=win32.constants.wdMove)
        #         doc.Range(sentence.Start, sentence.Start).InsertBefore(cell_value)
        #         break

        # 查找目标文本
        for paragraph in doc.Paragraphs:
            if object_str in paragraph.Range.Text:
                # 记录下最后一个目标文本的位置
                last_target_pos = paragraph.Range.Start

        # 在最后一个目标文本的前面插入字符串
        if last_target_pos > 0:
            doc.Range(last_target_pos, last_target_pos).InsertBefore(cell_value)

        # selection2.InsertBefore(cell_value)

    # 保存
    doc.SaveAs(os.path.join(result_dir, 'Word.doc'))
    print("修改完成！ 汇总签字表保存在: " + os.path.join(result_dir, 'Word.doc'))

    # 另存为PDF
    doc.SaveAs(os.path.join(result_dir, 'Word.pdf'), FileFormat=win32.constants.wdFormatPDF)
    print("打印完成！ 打印后的PDF保存在: " + os.path.join(result_dir, 'Word.pdf'))
    doc.Close()

    # 关闭Word文档和应用程序
    doc1.Close()

    word.Quit()
