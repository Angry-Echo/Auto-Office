import copy
import openpyxl
from openpyxl.utils import get_column_letter

path = r'C:\Users\xxx\Documents\test_dir - 副本\新建文件夹\新建文件夹\各科室汇总\开发一路2022年完成和在编的开发方案（CO2驱微观驱替规律与推广应用注采井网方式研究）承担情况统计表-采收率中心-俞宏伟.xlsx'
save_path = r'C:\Users\xxx\Documents\test_dir - 副本\新建文件夹\新建文件夹\各科室汇总\Results\new_file.xlsx'

wb = openpyxl.load_workbook(path)
wb2 = openpyxl.Workbook()

sheetnames = wb.sheetnames
for sheetname in sheetnames:
    print(sheetname)
    sheet = wb[sheetname]
    sheet2 = wb2.create_sheet(sheetname)

    # tab颜色
    sheet2.sheet_properties.tabColor = sheet.sheet_properties.tabColor

    # 开始处理合并单元格形式为“(<CellRange A1：A4>,)，替换掉(<CellRange 和 >,)' 找到合并单元格
    wm = list(sheet.merged_cells)
    if len(wm) > 0:
        for i in range(0, len(wm)):
            cell2 = str(wm[i]).replace('(<CellRange ', '').replace('>,)', '')
            sheet2.merge_cells(cell2)

    for i, row in enumerate(sheet.iter_rows()):
        sheet2.row_dimensions[i+1].height = sheet.row_dimensions[i+1].height
        for j, cell in enumerate(row):
            sheet2.column_dimensions[get_column_letter(j+1)].width = sheet.column_dimensions[get_column_letter(j+1)].width
            sheet2.cell(row=i + 1, column=j + 1, value=cell.value)

            # 设置单元格格式
            source_cell = sheet.cell(i+1, j+1)
            target_cell = sheet2.cell(i+1, j+1)
            target_cell.fill = copy.copy(source_cell.fill)
            if source_cell.has_style:
                target_cell._style = copy.copy(source_cell._style)
                target_cell.font = copy.copy(source_cell.font)
                target_cell.border = copy.copy(source_cell.border)
                target_cell.fill = copy.copy(source_cell.fill)
                target_cell.number_format = copy.copy(source_cell.number_format)
                target_cell.protection = copy.copy(source_cell.protection)
                target_cell.alignment = copy.copy(source_cell.alignment)

if 'Sheet' in wb2.sheetnames:
    del wb2['Sheet']
wb2.save(save_path)

wb.close()
wb2.close()

print('Done.')
