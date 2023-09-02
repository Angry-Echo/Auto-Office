import win32com.client as win32
import openpyxl
import copy
import glob
import os
import sys
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

import customtkinter
from customtkinter import filedialog

customtkinter.set_appearance_mode("System")  # Modes: "System" (standard), "Dark", "Light"
customtkinter.set_default_color_theme("blue")  # Themes: "blue" (standard), "green", "dark-blue"


def get_max_row(sheet, num_head_row):
    valid_value_position = num_head_row + 1

    while True:
        row_dict = {cell.value for cell in sheet[valid_value_position]}
        if row_dict != {None} and len(row_dict) > 2:
            valid_value_position += 1
        else:
            break

    return valid_value_position - 1


def Special_Find_Max_Head(root_dir):
    row_flag = 1
    try:
        excel_path = glob.glob(os.path.join(root_dir, '*.xlsx'))

        wb_1 = openpyxl.load_workbook(excel_path[0])
        wb_2 = openpyxl.load_workbook(excel_path[1])

        wb_1_sheet = wb_1.worksheets[0]
        wb_2_sheet = wb_2.worksheets[0]

        while True:
            wb_1_sheet_row = {cell.value for cell in wb_1_sheet[row_flag]}
            wb_2_sheet_row = {cell.value for cell in wb_2_sheet[row_flag]}

            if wb_1_sheet_row == wb_2_sheet_row:
                row_flag += 1
            else:
                break

    except PermissionError:
        print('有Excel文件未关闭！请关闭后重试')
        sys.exit()

    return row_flag - 1  # row_flag比最大行数多一个，是非表头行的位置


class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()

        # configure window
        self.title("自动化辅助高校办公系统——小鱼1号")
        self.geometry(f"{1100}x{580}")

        # configure grid layout (4x4)
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure((2, 3), weight=0)
        self.grid_rowconfigure((0, 1, 2), weight=1)

        # 外观边栏设置
        self.sidebar_frame = customtkinter.CTkFrame(self, width=140, corner_radius=0)
        self.sidebar_frame.grid(row=0, column=0, rowspan=6, sticky="nsew")
        self.sidebar_frame.grid_rowconfigure(6, weight=1)

        self.logo_label = customtkinter.CTkLabel(self.sidebar_frame, text="外观设置", font=customtkinter.CTkFont(family='华文楷体', size=30, weight="normal"))
        self.logo_label.grid(row=0, column=0, padx=20, pady=(20, 10))

        '''
        self.sidebar_button_1 = customtkinter.CTkButton(self.sidebar_frame, command=self.sidebar_button_event)
        self.sidebar_button_1.grid(row=1, column=0, padx=20, pady=10)

        self.sidebar_button_2 = customtkinter.CTkButton(self.sidebar_frame, command=self.sidebar_button_event)
        self.sidebar_button_2.grid(row=2, column=0, padx=20, pady=10)

        self.sidebar_button_3 = customtkinter.CTkButton(self.sidebar_frame, command=self.sidebar_button_event)
        self.sidebar_button_3.grid(row=3, column=0, padx=20, pady=10)
        '''

        self.appearance_mode_label = customtkinter.CTkLabel(self.sidebar_frame, font=customtkinter.CTkFont(family='楷体', size=20, weight="bold"), text="主题模式:", anchor="s")
        self.appearance_mode_label.grid(row=8, column=0, padx=20, pady=(10, 0))
        self.appearance_mode_optionemenu = customtkinter.CTkOptionMenu(self.sidebar_frame, font=customtkinter.CTkFont(family='楷体', size=15, weight="bold"),
                                                                       values=["光明", "黑暗", "跟随系统"], command=self.change_appearance_mode_event)
        self.appearance_mode_optionemenu.grid(row=9, column=0, padx=20, pady=(10, 10))

        self.scaling_label = customtkinter.CTkLabel(self.sidebar_frame, font=customtkinter.CTkFont(family='楷体', size=20, weight="bold"), text="界面缩放:", anchor="w")
        self.scaling_label.grid(row=10, column=0, padx=20, pady=(10, 0))

        self.scaling_optionemenu = customtkinter.CTkOptionMenu(self.sidebar_frame, values=["80%", "90%", "100%", "110%", "120%"], command=self.change_scaling_event)
        self.scaling_optionemenu.grid(row=11, column=0, padx=20, pady=(10, 20))

        # 操作说明
        self.textbox = customtkinter.CTkTextbox(self, font=customtkinter.CTkFont(family='华文楷体', size=25, weight="normal"), width=250)
        self.textbox.grid(row=0, column=1, padx=(20, 0), pady=(20, 0), sticky="nsew")

        # Excel目录路径输入
        self.entry_text = customtkinter.StringVar(self, value="待操作的Excel文件的路径")
        self.entry = customtkinter.CTkEntry(self, textvariable=self.entry_text, state='readonly')
        self.entry.grid(row=3, column=1, columnspan=2, padx=(20, 0), pady=(20, 10), sticky="nsew")

        self.main_button_1 = customtkinter.CTkButton(master=self, command=self.get_dir_path, text='选择路径', fg_color="transparent", border_width=2, text_color=("gray10", "#DCE4EE"))
        self.main_button_1.grid(row=3, column=3, padx=(20, 20), pady=(20, 10), sticky="nsew")

        # 签字表文件输入
        self.sign_entry_text = customtkinter.StringVar(self, value="示例的专家签字表文件")
        self.sign_entry = customtkinter.CTkEntry(self, textvariable=self.sign_entry_text, state='readonly')
        self.sign_entry.grid(row=5, column=1, columnspan=2, padx=(20, 0), pady=(0, 20), sticky="nsew")

        self.sign_button_1 = customtkinter.CTkButton(master=self, command=self.get_sign_file, text='选择文件', fg_color="transparent", border_width=2, text_color=("gray10", "#DCE4EE"))
        self.sign_button_1.grid(row=5, column=3, padx=(20, 20), pady=(0, 20), sticky="nsew")

        # 合并表设置
        self.tabview = customtkinter.CTkTabview(self, width=250)
        self.tabview.grid(row=0, column=2, padx=(20, 0), pady=(20, 0), sticky="nsew")
        self.tabview.add("表头")
        self.tabview.add("序号")
        self.tabview.add("字体")
        self.tabview.tab("表头").grid_columnconfigure(0, weight=1)  # configure grid of individual tabs
        self.tabview.tab("序号").grid_columnconfigure(0, weight=1)

        self.head_label = customtkinter.CTkLabel(self.tabview.tab("表头"), text="表头所在行", font=customtkinter.CTkFont(family='华文楷体', size=20, weight="normal"))
        self.head_label.grid(row=1, column=0, padx=20, pady=(10, 10))
        self.combobox_1 = customtkinter.CTkComboBox(self.tabview.tab("表头"), values=["1", "2", "3", "可以自己输入数字"])
        self.combobox_1.grid(row=2, column=0, padx=20, pady=(10, 10))

        self.id_label = customtkinter.CTkLabel(self.tabview.tab("序号"), text="序号所在列", font=customtkinter.CTkFont(family='华文楷体', size=20, weight="normal"))
        self.id_label.grid(row=1, column=0, padx=20, pady=(10, 10))
        self.combobox_2 = customtkinter.CTkComboBox(self.tabview.tab("序号"), values=["1", "2", "3", "可以自己输入数字"])
        self.combobox_2.grid(row=2, column=0, padx=20, pady=(10, 10))

        self.font_label = customtkinter.CTkLabel(self.tabview.tab("字体"), text="字号（14）字体（仿宋）", font=customtkinter.CTkFont(family='华文楷体', size=18, weight="normal"))
        self.font_label.grid(row=1, column=0, padx=20, pady=(10, 10))
        self.switch = customtkinter.CTkSwitch(self.tabview.tab("字体"), text="自动修改字体")
        self.switch.grid(row=2, column=0, padx=10, pady=(0, 20))

        '''
        self.optionmenu_1 = customtkinter.CTkOptionMenu(self.tabview.tab("表头"), dynamic_resizing=False, values=["Value 1", "Value 2", "Value Long Long Long"])
        self.optionmenu_1.grid(row=2, column=0, padx=20, pady=(20, 10))
        self.string_input_button = customtkinter.CTkButton(self.tabview.tab("表头"), text="Open CTkInputDialog", command=self.open_input_dialog_event)
        self.string_input_button.grid(row=3, column=0, padx=20, pady=(10, 10))
        self.label_tab_2 = customtkinter.CTkLabel(self.tabview.tab("序号"), text="CTkLabel on Tab 2")
        self.label_tab_2.grid(row=0, column=0, padx=20, pady=20)
        '''

        # 合并表运行
        self.radiobutton_frame = customtkinter.CTkFrame(self)
        self.radiobutton_frame.grid(row=0, column=3, padx=(20, 20), pady=(40, 0), sticky="nsew")

        self.label_radio_group = customtkinter.CTkLabel(master=self.radiobutton_frame, font=customtkinter.CTkFont(family='华文楷体', size=20, weight="bold"), text="合并项目表")
        self.label_radio_group.grid(row=1, column=1, columnspan=1, padx=10, pady=10, sticky="")
        '''
        self.radio_var = tkinter.IntVar(value=0)
        self.radio_button_1 = customtkinter.CTkRadioButton(master=self.radiobutton_frame, variable=self.radio_var, value=0)
        self.radio_button_1.grid(row=1, column=2, pady=10, padx=20, sticky="n")
        self.radio_button_2 = customtkinter.CTkRadioButton(master=self.radiobutton_frame, variable=self.radio_var, value=1)
        self.radio_button_2.grid(row=2, column=2, pady=10, padx=20, sticky="n")
        self.radio_button_3 = customtkinter.CTkRadioButton(master=self.radiobutton_frame, variable=self.radio_var, value=2)
        self.radio_button_3.grid(row=3, column=2, pady=10, padx=20, sticky="n")
        '''
        self.Combine = customtkinter.CTkButton(master=self.radiobutton_frame, command=self.Merge_Excel, font=customtkinter.CTkFont(size=20, weight="bold"), text='运行',
                                               fg_color="transparent", border_width=2, text_color=("gray10", "#DCE4EE"))
        self.Combine.grid(row=3, column=1, padx=(20, 20), pady=(33, 20), sticky="nsew")

        # 日志记录仪
        self.slider_progressbar_frame = customtkinter.CTkFrame(self, fg_color="transparent")
        self.slider_progressbar_frame.grid(row=1, column=1, padx=(20, 0), pady=(20, 0), sticky="nsew")
        self.slider_progressbar_frame.grid_columnconfigure(0, weight=1)
        self.slider_progressbar_frame.grid_rowconfigure(4, weight=1)

        self.seg_button_1 = customtkinter.CTkSegmentedButton(self.slider_progressbar_frame)
        self.seg_button_1.grid(row=0, column=0, padx=(20, 10), pady=(10, 10), sticky="ew")
        self.progressbar_1 = customtkinter.CTkProgressBar(self.slider_progressbar_frame)
        self.progressbar_1.grid(row=1, column=0, padx=(20, 10), pady=(10, 10), sticky="ew")

        self.recorder = customtkinter.CTkLabel(master=self.slider_progressbar_frame, font=customtkinter.CTkFont(family='华文楷体', size=20, weight="bold"), text="暂无执行动作")
        self.recorder.grid(row=2, column=0, columnspan=1, padx=10, pady=10, sticky="")

        # 签字表设置
        self.tabview_2 = customtkinter.CTkTabview(self, width=250)
        self.tabview_2.grid(row=1, column=2, padx=(20, 0), pady=(20, 0), sticky="nsew")
        self.tabview_2.add("标题")
        self.tabview_2.add("列号")
        self.tabview_2.tab("标题").grid_columnconfigure(0, weight=1)  # configure grid of individual tabs
        self.tabview_2.tab("列号").grid_columnconfigure(0, weight=1)

        self.detection_head_label = customtkinter.CTkLabel(self.tabview_2.tab("标题"), text="签字表标题", font=customtkinter.CTkFont(family='华文楷体', size=20, weight="normal"))
        self.detection_head_label.grid(row=1, column=0, padx=20, pady=(10, 10))
        self.head_entry = customtkinter.CTkEntry(self.tabview_2.tab("标题"), placeholder_text="要检测的标题")
        self.head_entry.grid(row=2, column=0, padx=20, pady=(10, 10))

        self.proj_id_label = customtkinter.CTkLabel(self.tabview_2.tab("列号"), text="项目所在列", font=customtkinter.CTkFont(family='华文楷体', size=20, weight="normal"))
        self.proj_id_label.grid(row=1, column=0, padx=20, pady=(10, 10))
        self.proj_combobox = customtkinter.CTkComboBox(self.tabview_2.tab("列号"), values=["A", "B", "C", "可以自己输入字母"])
        self.proj_combobox.grid(row=2, column=0, padx=20, pady=(10, 10))

        # 签字表运行
        self.sign_frame = customtkinter.CTkFrame(self)
        self.sign_frame.grid(row=1, column=3, padx=(20, 20), pady=(40, 0), sticky="nsew")
        self.sign_label = customtkinter.CTkLabel(master=self.sign_frame, font=customtkinter.CTkFont(family='华文楷体', size=20, weight="bold"), text="生成签字表")
        self.sign_label.grid(row=1, column=1, columnspan=1, padx=10, pady=10, sticky="")

        self.sign_Combine = customtkinter.CTkButton(master=self.sign_frame, command=self.Word_PDF, font=customtkinter.CTkFont(size=20, weight="bold"), text='运行',
                                                    fg_color="transparent", border_width=2, text_color=("gray10", "#DCE4EE"))
        self.sign_Combine.grid(row=3, column=1, padx=(20, 20), pady=(33, 20), sticky="nsew")

        # set default values
        '''self.sidebar_button_3.configure(state="disabled", text="Disabled CTkButton")'''
        # self.checkbox_3.configure(state="disabled")
        # self.checkbox_1.select()
        # self.scrollable_frame_switches[0].select()
        # self.scrollable_frame_switches[4].select()
        '''self.radio_button_3.configure(state="disabled")'''
        self.appearance_mode_optionemenu.set("黑暗")
        self.scaling_optionemenu.set("100%")
        '''self.optionmenu_1.set("CTkOptionmenu")'''
        self.combobox_1.set("1")
        # self.slider_1.configure(command=self.progressbar_2.set)
        # self.slider_2.configure(command=self.progressbar_3.set)
        self.progressbar_1.configure(mode="indeterminnate")
        self.progressbar_1.start()
        self.textbox.insert("0.0", "使用手册：\n\n")
        self.seg_button_1.configure(values=["运行日志记录仪", "运动模式"])
        self.seg_button_1.set("运行日志记录仪")

    def open_input_dialog_event(self):
        dialog = customtkinter.CTkInputDialog(text="Type in a number:", title="CTkInputDialog")
        print("CTkInputDialog:", dialog.get_input())

    def change_appearance_mode_event(self, new_appearance_mode: str):
        customtkinter.set_appearance_mode(new_appearance_mode)

    def change_scaling_event(self, new_scaling: str):
        new_scaling_float = int(new_scaling.replace("%", "")) / 100
        customtkinter.set_widget_scaling(new_scaling_float)

    def sidebar_button_event(self):
        print("sidebar_button click")

    def get_dir_path(self):
        path = filedialog.askdirectory(title='请选择要合并的Excel文件所在目录')  # 返回一个字符串，且只能获取文件夹路径，不能获取文件的路径
        self.entry_text.set(path)

    def get_sign_file(self):
        sign_path = filedialog.askopenfilename(title='请选择示例的专家签字表Word文件')  # 返回一个字符串，且只能获取文件的路径
        self.sign_entry_text.set(sign_path)

    def Merge_Excel(self):
        root_dir = self.entry_text.get()
        head_row = int(self.combobox_1.get())
        col_id = int(self.combobox_2.get())
        font_flag = int(self.switch.get())

        # 创建一个新的工作表
        merged = openpyxl.Workbook()
        merged_sheet = merged.active

        # 遍历所有excel文件的sheet,存为list
        sheet_list = []
        try:
            for path in glob.glob(os.path.join(root_dir, '*.xlsx')):  # 搜索给定目录下所有的.xlsx文件
                wb = openpyxl.load_workbook(path)
                sheet = wb.worksheets[0]
                sheet_list.append(sheet)
        except PermissionError:
            self.recorder.configure(text='合并失败，这是因为目录中有Excel文件未关闭！请关闭后重试')
            sys.exit()

        '''
        # 该目录下Excel文件的共同表头的行数
        head_row = Find_Max_Head(sheet_list)
        '''

        # 合并所有sheet中的数据，带格式，复制到新的工作表中
        for n_s, sheet in enumerate(sheet_list):
            num_whole_row = get_max_row(sheet, head_row)

            if n_s == 0:
                place_holder = 0  # 第一个表的表头需要，所有没有行是无效的
                row_begin = 0
            elif n_s == 1:
                place_holder = head_row  # 第二个表的表头不需要，有几行是无效的
                last_max_row = get_max_row(sheet_list[n_s - 1], head_row)  # 现在是第二个表，它的前一个表是第一张表（也就是算入了表头）
                row_begin += last_max_row
            else:
                place_holder = head_row  # 同上
                last_max_row = get_max_row(sheet_list[n_s - 1], head_row) - head_row  # 现在是第三个表，它的前一个表是第二个表（行数不计入表头
                row_begin += last_max_row

            # tab颜色
            merged_sheet.sheet_properties.tabColor = sheet.sheet_properties.tabColor

            # 开始处理合并单元格形式为“(<CellRange A1：A4>,)，替换掉(<CellRange 和 >,)' 找到合并单元格
            wm = list(sheet.merged_cells)
            if len(wm) > 0:
                for i in range(0, len(wm)):
                    cell_2 = str(wm[i]).replace('(<CellRange ', '').replace('>,)', '')
                    merged_sheet.merge_cells(cell_2)

            for n_r in range(num_whole_row):
                # 设定行高
                merged_sheet.row_dimensions[n_r + 1].height = sheet.row_dimensions[n_r + 1].height

                row = sheet[n_r + place_holder + 1]  # excel 坐标从1开始

                for n_c, source_cell in enumerate(row):
                    # 设定列宽
                    merged_sheet.column_dimensions[get_column_letter(n_c + 1)].width = sheet.column_dimensions[get_column_letter(n_c + 1)].width

                    # 复制单元格，带格式
                    target_cell = merged_sheet.cell(row=row_begin + n_r + 1, column=n_c + 1)
                    try:
                        target_cell.value = source_cell.value
                    except AttributeError:
                        pass  # 合并后只要写第一个即可，后面的都被合并进第一个了，当然找不到了

                    target_cell.fill = copy.copy(source_cell.fill)

                    if source_cell.has_style:
                        target_cell._style = copy.copy(source_cell._style)
                        target_cell.font = copy.copy(source_cell.font)
                        target_cell.border = copy.copy(source_cell.border)
                        target_cell.fill = copy.copy(source_cell.fill)
                        target_cell.number_format = copy.copy(source_cell.number_format)
                        target_cell.protection = copy.copy(source_cell.protection)
                        target_cell.alignment = copy.copy(source_cell.alignment)

        # 从第二行开始遍历，将 "序号" 列中的值递增
        merged_max_row = get_max_row(merged_sheet, head_row)

        for row in range(head_row + 1, merged_max_row + 1):
            merged_sheet.cell(row=row, column=col_id).value = row - head_row  # 妙！

        # 是否修改字体
        if font_flag == 1:
            for row in range(head_row + 1, merged_max_row + 1):
                sheet_row = merged_sheet[row]
                for cell in sheet_row:
                    font = Font(name="仿宋", size=14)
                    cell.font = font

        # 创建结算文件夹，保存新的Excel文件
        try:
            if not os.path.exists(os.path.join(root_dir, 'Results')):
                os.mkdir(os.path.join(root_dir, 'Results'))
        except FileNotFoundError:
            self.recorder.configure(text='您还没有选择相应的路径或文件！')

        merged.save(os.path.join(root_dir, 'Results', 'Merged.xlsx'))
        self.recorder.configure(text="合并完成! \n 汇总项目表保存在: \n" + root_dir + '/' + '\n' + 'Results/' + '\n' + 'Merged.xlsx' + '\n')

    def Word_PDF(self):

        try:
            result_dir = os.path.join(self.entry_text.get(), 'Results')
        except FileNotFoundError:
            self.recorder.configure(text="没有成功生成合并表 \n 所以无法生成签字表")

        example_path = self.sign_entry_text.get()
        object_str = self.head_entry.get()
        porj_col_id = self.proj_combobox.get()
        head_row = int(self.combobox_1.get())

        # 创建Word应用程序对象
        word = win32.gencache.EnsureDispatch('Word.Application')
        word.Visible = True

        # 打开Excel文件并读取数据
        try:
            wb = openpyxl.load_workbook(os.path.join(result_dir, 'Merged.xlsx'))
            ws = wb.active
        except PermissionError:
            self.recorder.configure(text='打印失败，这是因为合并表Excel文件未关闭！请关闭后重试')
            sys.exit()

        # 从哪一行开始获取项目名称：共同表头的下一行
        proj_start_row = head_row + 1

        # 打开要复制内容的Word文档
        doc1 = word.Documents.Open(example_path)

        # 选择要复制的内容
        selection = word.Selection
        selection.WholeStory()
        selection.Copy()

        # 将内容粘贴到新打开的Word文档中
        doc = word.Documents.Add()
        selection2 = word.Selection
        selection2.EndKey(win32.constants.wdStory)
        selection2.Paste()

        selection2.EndKey(win32.constants.wdStory)
        selection2.InsertBreak(win32.constants.wdPageBreak)

        # 获取文字
        first_cell_value = ws[porj_col_id][proj_start_row - 1].value  # 索引是从0开始，所以 position - 1
        try:
            if '《' in first_cell_value:
                pass
            else:
                first_cell_value = '《' + first_cell_value + '》'
        except TypeError:
            self.recorder.configure(text='您没有正确选择项目列号！\n 大概率是您选择了序号所在列')  # 数字不可迭代，这是选择了序号所在的列（很可能是没更改，一般默认的A就是序号列）

        # selection2.Find.Execute('开放性课题', False, False, False, False, False, True, 1, False, first_cell_value, 2)

        # selection2.EndKey(win32.constants.wdStory)
        # selection2.MoveUp(Unit=win32.constants.wdLine, Count=10)
        # selection2.HomeKey(Unit=win32.constants.wdLine)
        # selection2.InsertBefore(first_cell_value)

        # for i in range(10):
        #     selection2.Delete()

        # 查找目标句子
        for sentence in doc.Sentences:
            if object_str in sentence.Text:
                # 找到目标句子后，在其前面插入字符串
                sentence.StartOf(Unit=win32.constants.wdSentence, Extend=win32.constants.wdMove)
                doc.Range(sentence.Start, sentence.Start).InsertBefore(first_cell_value)
                break

        data = [cell.value for cell in ws[porj_col_id][proj_start_row - 1 + 1:] if cell.value is not None]  # 假设需要替换的文本在第一列

        last_target_pos = 0
        for cell_value in data:
            if '《' in cell_value:
                pass
            else:
                cell_value = '《' + cell_value + '》'

            selection2 = word.Selection
            selection2.EndKey(win32.constants.wdStory)
            selection2.Paste()

            selection2.EndKey(win32.constants.wdStory)
            selection2.InsertBreak(win32.constants.wdPageBreak)

            # selection2.Find.Execute('开放性课题', False, False, False, False, False, True, 1, False, cell_value, 2)

            # selection2.EndKey(win32.constants.wdStory)
            # selection2.MoveUp(Unit=win32.constants.wdLine, Count=10)
            # selection2.HomeKey(Unit=win32.constants.wdLine)

            # for i in range(10):
            #     selection2.Delete()

            # 查找目标句子
            # target_text = object_str
            # for sentence in doc.Sentences:
            #     if target_text in sentence.Text:
            #         # 找到目标句子后，在其前面插入字符串
            #         sentence.StartOf(Unit=win32.constants.wdSentence, Extend=win32.constants.wdMove)
            #         doc.Range(sentence.Start, sentence.Start).InsertBefore(cell_value)
            #         break

            # 查找目标文本
            for paragraph in doc.Paragraphs:
                if object_str in paragraph.Range.Text:
                    # 记录下最后一个目标文本的位置
                    last_target_pos = paragraph.Range.Start

            # 在最后一个目标文本的前面插入字符串
            if last_target_pos > 0:
                doc.Range(last_target_pos, last_target_pos).InsertBefore(cell_value)

            # selection2.InsertBefore(cell_value)

        # 保存
        doc.SaveAs(os.path.join(result_dir, 'Word.doc'))
        self.recorder.configure(text="修改完成！\n 汇总签字表保存在:\n" + result_dir + '\n' + 'Word.doc')

        # 另存为PDF
        doc.SaveAs(os.path.join(result_dir, 'Word.pdf'), FileFormat=win32.constants.wdFormatPDF)
        self.recorder.configure(text="打印完成！\n 汇总签字表PDF保存在:\n" + result_dir + '\n' + 'Word.pdf')
        doc.Close()

        # 关闭Word文档和应用程序
        doc1.Close()

        word.Quit()


if __name__ == "__main__":
    app = App()
    app.mainloop()
    print(app.entry_text.get())
