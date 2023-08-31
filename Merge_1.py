# -*- coding: utf-8 -*-
import openpyxl
import copy
import glob
import os

# 修改目录work_dir
work_dir = r'./test_dir'

# 创建一个新的工作表
merged = openpyxl.Workbook()
merged_sheet = merged.active


def get_max_row(sheet):
    real_max_row = 1

    while True:
        row_dict = {cell.value for cell in sheet[real_max_row]}
        if row_dict != {None}:
            real_max_row += 1
        else:
            break

    return real_max_row - 1


# 遍历所有excel文件的sheet,存为list
sheet_list = []
for path in glob.glob(os.path.join(work_dir, '*.xlsx')):  # 搜索给定目录下所有的.xlsx文件
    wb = openpyxl.load_workbook(path)
    sheet = wb.get_sheet_by_name('Sheet1')
    sheet_list.append(sheet)

# 合并所有sheet中的数据，带格式，复制到新的工作表中
for n_s, sheet in enumerate(sheet_list):
    current_max_row = get_max_row(sheet)

    if n_s == 0:

        row_begin = 0
    else:

        last_max_row = get_max_row(sheet_list[n_s - 1])
        row_begin += last_max_row

    for n_r in range(current_max_row):
        row = sheet[n_r + 1]  # excel 坐标从1开始

        for n_c, source_cell in enumerate(row):

            # 复制单元格，带格式
            target_cell = merged_sheet.cell(row=row_begin + n_r + 1, column=n_c + 1)
            target_cell.value = source_cell.value
            target_cell.fill = copy.copy(source_cell.fill)

            if source_cell.has_style:
                target_cell._style = copy.copy(source_cell._style)
                target_cell.font = copy.copy(source_cell.font)
                target_cell.border = copy.copy(source_cell.border)
                target_cell.fill = copy.copy(source_cell.fill)
                target_cell.number_format = copy.copy(source_cell.number_format)
                target_cell.protection = copy.copy(source_cell.protection)
                target_cell.alignment = copy.copy(source_cell.alignment)

# 保存新的Excel文件
merged.save(os.path.join(work_dir, 'merged2.xlsx'))
print("save excel to: " + os.path.join(work_dir, 'merged'))
