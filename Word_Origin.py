import win32com.client as win32
import openpyxl
import os


def Word_PDF(root_dir, example_path):
    # 创建Word应用程序对象
    word = win32.Dispatch('Word.Application')
    word.Visible = True

    # 打开Excel文件并读取数据
    wb = openpyxl.load_workbook(os.path.join(root_dir, 'Merged.xlsx'))
    ws = wb.active

    # 打开要复制内容的Word文档
    doc1 = word.Documents.Open(example_path)

    # 选择要复制的内容
    selection = word.Selection
    selection.WholeStory()
    selection.Copy()

    # 将内容粘贴到新打开的Word文档中
    doc = word.Documents.Add()
    selection2 = word.Selection
    selection2.EndKey(win32.constants.wdStory)
    selection2.Paste()

    # 替换文字
    first_cell_value = ws['B'][1].value
    if '《' in first_cell_value:
        pass
    else:
        first_cell_value = '《' + first_cell_value + '》'

    selection2.Find.Execute('开放性课题', False, False, False, False, False, True, 1, False, first_cell_value, 2)

    selection2.EndKey(win32.constants.wdStory)
    selection2.MoveUp(Unit=win32.constants.wdLine, Count=2)
    selection2.HomeKey(Unit=win32.constants.wdLine)
    for i in range(10):
        selection2.Delete()

    selection2.EndKey(win32.constants.wdStory)
    selection2.InsertBreak(win32.constants.wdPageBreak)

    # 保存新的Word文档
    doc.SaveAs(os.path.join(root_dir, 'Word.doc'))
    print("Finish! Saving Word to: " + os.path.join(root_dir, 'Word.doc'))

    data = [cell.value for cell in ws['B'][2:]]  # 假设需要替换的文本在第一列
    for cell_value in data:
        if cell_value != '外协项目名称':
            if '《' in cell_value:
                pass
            else:
                cell_value = '《' + cell_value + '》'

            selection2 = word.Selection
            selection2.EndKey(win32.constants.wdStory)
            selection2.Paste()

            selection2.Find.Execute('开放性课题', False, False, False, False, False, True, 1, False, cell_value, 2)

            selection2.EndKey(win32.constants.wdStory)
            selection2.MoveUp(Unit=win32.constants.wdLine, Count=2)
            selection2.HomeKey(Unit=win32.constants.wdLine)
            for i in range(10):
                selection2.Delete()

            selection2.EndKey(win32.constants.wdStory)
            selection2.InsertBreak(win32.constants.wdPageBreak)

        else:
            pass

    doc.SaveAs(os.path.join(root_dir, 'Word.pdf'), FileFormat=win32.constants.wdFormatPDF)
    print("Finish! Saving PDF to: " + os.path.join(root_dir, 'Word.pdf'))

    # 关闭Word文档和应用程序
    doc1.Close()
    doc.Close()

    word.Quit()
